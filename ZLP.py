from cvxopt.modeling import variable, op
from openpyxl import load_workbook

# Открытие файла
workbook = load_workbook(filename="C:/Users/User/Desktop/Постановка курсача.xlsx")

# Считывание листа
sheet = workbook["Лист1"]

# Списки
Stock = [] #запасы
Needs = [] #потребности
Ways = [] #цены за перевозку

# Необходимые переменные
SumStock = 0
SumNeeds = 0

# заполнение запасов
for i in range(2, 12): # Строки
    cell = sheet.cell(row = i, column = 12).value
    Stock.append(cell)
    SumStock += cell

# заполнение потребностей
for i in range(2, 12): # Столбцы
    cell = sheet.cell(row = 12, column = i).value
    Needs.append(cell)
    SumNeeds += cell

# заполнение путей
for i in range(2, 12):
    for j in range(2, 12):
        cell = sheet.cell(row = i, column = j).value
        Ways.append(cell)

# Вывод начальных данных
print("=========================================================================================================================")
print("Потребности (тонны): ")
needs = ""
for i in range(len(Needs)):
    needs += str(Needs[i]) + " "
print(needs)

print("Запасы (тонны): ")
stock = ""
for i in range(len(Stock)):
    stock += str(Stock[i]) + " "
print(stock)

print("Стоимость доставки (рубли): ")
counter = 0
ways = ""
for i in range(len(Ways)):
    counter += 1
    if (counter % 10 == 0):
        ways += str(Ways[i]).ljust(5, " ") + "\n"
    else:
        ways += str(Ways[i]).ljust(5, " ") + " | "
print(ways)

# Определение типа задачи
print("=========================================================================================================================")
print("Тип задачи: ")

type = 0 # 1 - открытая, 2 - открытая, 3 - сбалансированная

if (SumNeeds == SumStock):
    type = 1
    print("Задача закрыта - спрос равен предложению")
elif (SumNeeds >= SumStock):
    type = 2
    print("Задача открытого типа - спрос превышает предложение")
elif (SumNeeds <= SumStock):
    type = 3
    print("Задача открытого типа - предложение превышает спрос")
print("=========================================================================================================================")
x = variable(100, 'x') # объявляем искомые переменные

# Заполнение целевой функции
z = 0
for i in range (100):
    z += Ways[i] * x[i]

# Заполнение системы ограничений
if type == 3: # Зактытый тип
    # Строки
    mass1 = (x[0] + x[1] + x[2] + x[3] + x[4] + x[5] + x[6] + x[7] + x[8] + x[9] <= Stock[0])
    mass2 = (x[10] + x[11] + x[12] + x[13] + x[14] + x[15] + x[16] + x[17] + x[18] + x[19] <= Stock[1])
    mass3 = (x[20] + x[21] + x[22] + x[23] + x[24] + x[25] + x[26] + x[27] + x[28] + x[29] <= Stock[2])
    mass4 = (x[30] + x[31] + x[32] + x[33] + x[34] + x[35] + x[36] + x[37] + x[38] + x[39] <= Stock[3])
    mass5 = (x[40] + x[41] + x[42] + x[43] + x[44] + x[45] + x[46] + x[47] + x[48] + x[49] <= Stock[4])
    mass6 = (x[50] + x[51] + x[52] + x[53] + x[54] + x[55] + x[56] + x[57] + x[58] + x[59] <= Stock[5])
    mass7 = (x[60] + x[61] + x[62] + x[63] + x[64] + x[65] + x[66] + x[67] + x[68] + x[69] <= Stock[6])
    mass8 = (x[70] + x[71] + x[72] + x[73] + x[74] + x[75] + x[76] + x[77] + x[78] + x[79] <= Stock[7])
    mass9 = (x[80] + x[81] + x[82] + x[83] + x[84] + x[85] + x[86] + x[87] + x[88] + x[89] <= Stock[8])
    mass10 = (x[90] + x[91] + x[92] + x[93] + x[94] + x[95] + x[96] + x[97] + x[98] + x[99] <= Stock[9])

    # Столбцы
    mass11 = (x[0] + x[10] + x[20] + x[30] + x[40] + x[50] + x[60] + x[70] + x[80] + x[90] == Needs[0])
    mass12 = (x[1] + x[11] + x[21] + x[31] + x[41] + x[51] + x[61] + x[71] + x[81] + x[91] == Needs[1])
    mass13 = (x[2] + x[12] + x[22] + x[32] + x[42] + x[52] + x[62] + x[72] + x[82] + x[92] == Needs[2])
    mass14 = (x[3] + x[13] + x[23] + x[33] + x[43] + x[53] + x[63] + x[73] + x[83] + x[93] == Needs[3])
    mass15 = (x[4] + x[14] + x[24] + x[34] + x[44] + x[54] + x[64] + x[74] + x[84] + x[94] == Needs[4])
    mass16 = (x[5] + x[15] + x[25] + x[35] + x[45] + x[55] + x[65] + x[75] + x[85] + x[95] == Needs[5])
    mass17 = (x[6] + x[16] + x[26] + x[36] + x[46] + x[56] + x[66] + x[76] + x[86] + x[96] == Needs[6])
    mass18 = (x[7] + x[17] + x[27] + x[37] + x[47] + x[57] + x[67] + x[77] + x[87] + x[97] == Needs[7])
    mass19 = (x[8] + x[18] + x[28] + x[38] + x[48] + x[58] + x[68] + x[78] + x[88] + x[98] == Needs[8])
    mass20 = (x[9] + x[19] + x[29] + x[39] + x[49] + x[59] + x[69] + x[79] + x[89] + x[99] == Needs[9])
elif type == 2: # Открытый тип
    # Строки
    mass1 = (x[0] + x[1] + x[2] + x[3] + x[4] + x[5] + x[6] + x[7] + x[8] + x[9] <= Stock[0])
    mass2 = (x[10] + x[11] + x[12] + x[13] + x[14] + x[15] + x[16] + x[17] + x[18] + x[19] >= Stock[1])
    mass3 = (x[20] + x[21] + x[22] + x[23] + x[24] + x[25] + x[26] + x[27] + x[28] + x[29] >= Stock[2])
    mass4 = (x[30] + x[31] + x[32] + x[33] + x[34] + x[35] + x[36] + x[37] + x[38] + x[39] >= Stock[3])
    mass5 = (x[40] + x[41] + x[42] + x[43] + x[44] + x[45] + x[46] + x[47] + x[48] + x[49] >= Stock[4])
    mass6 = (x[50] + x[51] + x[52] + x[53] + x[54] + x[55] + x[56] + x[57] + x[58] + x[59] >= Stock[5])
    mass7 = (x[60] + x[61] + x[62] + x[63] + x[64] + x[65] + x[66] + x[67] + x[68] + x[69] >= Stock[6])
    mass8 = (x[70] + x[71] + x[72] + x[73] + x[74] + x[75] + x[76] + x[77] + x[78] + x[79] >= Stock[7])
    mass9 = (x[80] + x[81] + x[82] + x[83] + x[84] + x[85] + x[86] + x[87] + x[88] + x[89] >= Stock[8])
    mass10 = (x[90] + x[91] + x[92] + x[93] + x[94] + x[95] + x[96] + x[97] + x[98] + x[99] >= Stock[9])

    # Столбцы
    mass11 = (x[0] + x[10] + x[20] + x[30] + x[40] + x[50] + x[60] + x[70] + x[80] + x[90] == Needs[0])
    mass12 = (x[1] + x[11] + x[21] + x[31] + x[41] + x[51] + x[61] + x[71] + x[81] + x[91] == Needs[1])
    mass13 = (x[2] + x[12] + x[22] + x[32] + x[42] + x[52] + x[62] + x[72] + x[82] + x[92] == Needs[2])
    mass14 = (x[3] + x[13] + x[23] + x[33] + x[43] + x[53] + x[63] + x[73] + x[83] + x[93] == Needs[3])
    mass15 = (x[4] + x[14] + x[24] + x[34] + x[44] + x[54] + x[64] + x[74] + x[84] + x[94] == Needs[4])
    mass16 = (x[5] + x[15] + x[25] + x[35] + x[45] + x[55] + x[65] + x[75] + x[85] + x[95] == Needs[5])
    mass17 = (x[6] + x[16] + x[26] + x[36] + x[46] + x[56] + x[66] + x[76] + x[86] + x[96] == Needs[6])
    mass18 = (x[7] + x[17] + x[27] + x[37] + x[47] + x[57] + x[67] + x[77] + x[87] + x[97] == Needs[7])
    mass19 = (x[8] + x[18] + x[28] + x[38] + x[48] + x[58] + x[68] + x[78] + x[88] + x[98] == Needs[8])
    mass20 = (x[9] + x[19] + x[29] + x[39] + x[49] + x[59] + x[69] + x[79] + x[89] + x[99] == Needs[9])
elif type == 1: # Закрытая
    # Строки
    mass1 = (x[0] + x[1] + x[2] + x[3] + x[4] + x[5] + x[6] + x[7] + x[8] + x[9] <= Stock[0])
    mass2 = (x[10] + x[11] + x[12] + x[13] + x[14] + x[15] + x[16] + x[17] + x[18] + x[19] == Stock[1])
    mass3 = (x[20] + x[21] + x[22] + x[23] + x[24] + x[25] + x[26] + x[27] + x[28] + x[29] == Stock[2])
    mass4 = (x[30] + x[31] + x[32] + x[33] + x[34] + x[35] + x[36] + x[37] + x[38] + x[39] == Stock[3])
    mass5 = (x[40] + x[41] + x[42] + x[43] + x[44] + x[45] + x[46] + x[47] + x[48] + x[49] == Stock[4])
    mass6 = (x[50] + x[51] + x[52] + x[53] + x[54] + x[55] + x[56] + x[57] + x[58] + x[59] == Stock[5])
    mass7 = (x[60] + x[61] + x[62] + x[63] + x[64] + x[65] + x[66] + x[67] + x[68] + x[69] == Stock[6])
    mass8 = (x[70] + x[71] + x[72] + x[73] + x[74] + x[75] + x[76] + x[77] + x[78] + x[79] == Stock[7])
    mass9 = (x[80] + x[81] + x[82] + x[83] + x[84] + x[85] + x[86] + x[87] + x[88] + x[89] == Stock[8])
    mass10 = (x[90] + x[91] + x[92] + x[93] + x[94] + x[95] + x[96] + x[97] + x[98] + x[99] == Stock[9])

    # Столбцы
    mass11 = (x[0] + x[10] + x[20] + x[30] + x[40] + x[50] + x[60] + x[70] + x[80] + x[90] == Needs[0])
    mass12 = (x[1] + x[11] + x[21] + x[31] + x[41] + x[51] + x[61] + x[71] + x[81] + x[91] == Needs[1])
    mass13 = (x[2] + x[12] + x[22] + x[32] + x[42] + x[52] + x[62] + x[72] + x[82] + x[92] == Needs[2])
    mass14 = (x[3] + x[13] + x[23] + x[33] + x[43] + x[53] + x[63] + x[73] + x[83] + x[93] == Needs[3])
    mass15 = (x[4] + x[14] + x[24] + x[34] + x[44] + x[54] + x[64] + x[74] + x[84] + x[94] == Needs[4])
    mass16 = (x[5] + x[15] + x[25] + x[35] + x[45] + x[55] + x[65] + x[75] + x[85] + x[95] == Needs[5])
    mass17 = (x[6] + x[16] + x[26] + x[36] + x[46] + x[56] + x[66] + x[76] + x[86] + x[96] == Needs[6])
    mass18 = (x[7] + x[17] + x[27] + x[37] + x[47] + x[57] + x[67] + x[77] + x[87] + x[97] == Needs[7])
    mass19 = (x[8] + x[18] + x[28] + x[38] + x[48] + x[58] + x[68] + x[78] + x[88] + x[98] == Needs[8])
    mass20 = (x[9] + x[19] + x[29] + x[39] + x[49] + x[59] + x[69] + x[79] + x[89] + x[99] == Needs[9])

# x >= 0
x_non_negative = (x >= 0) 

# Заполнение системы ограничений
problem = op (z, [mass1, mass2, mass3, mass4, mass5, mass6, mass7, mass8, mass9, mass10, 
                  mass11, mass12, mass13, mass14, mass15, mass16, mass17, mass18, mass19, mass20, 
                  x_non_negative])
problem.solve(solver = 'glpk')

# Чтение названий и запись в массивы
Careers = [] # карьеры
Plots = [] # участки дороги

# заполнение названий карьеров
for i in range(2, 12): # Строки
    cell = sheet.cell(row = i, column = 1).value
    Careers.append(cell)

# заполнение названий участков
for i in range(2, 12): # Столбцы
    cell = sheet.cell(row = 1, column = i).value
    Plots.append(cell)

# Вывод результата
print("=========================================================================================================================")
print("План перевозок:")

columns = ""
for i in range(len(Plots)):
    columns += "| " + Plots[i] + "     "
print("                               " + columns)
    
counter = 0
system = ""
counter_careers = 0
for i in x.value:
    counter += 1
    if (counter % 10 == 0):
        system += str(i).ljust(6, " ") +  "\n" 
    elif (counter == 1 or counter % 10 == 1):
        system += str(Careers[counter_careers]).ljust(30, " ") + " | " + str(i).ljust(6, " ") + " | " 
        counter_careers += 1
    else:
        system += str(i).ljust(6, " ") +  " | "
print(system)
print("=========================================================================================================================")

# Интерпретация результата
print("Интерпретация плана перевозок:")
result = ""
counter = 0
counter_careers = 0
for i in x.value:
    counter += 1
    if (counter == 1 or counter % 10 == 1):
        result += "Из карьера " + Careers[counter_careers] + " требуется перевезти "
        counter_careers += 1
    if (counter % 10 == 0 and i != 0):
        result += "на участок " + Plots[9] + " - " + str(i) + " т "
    if (counter % 10 == 0):
        result += "\n"
    if (counter % 10 != 0 and i != 0):
        result += "на участок " + Plots[counter % 10 - 1] + " - " + str(i) + " т "
print(result)
print("=========================================================================================================================")

print("Минимальная стоимость перевозок (рубли):")
print(problem.objective.value()[0])
print("=========================================================================================================================")
